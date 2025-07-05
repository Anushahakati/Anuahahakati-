import face_recognition
import cv2
import numpy as np
from datetime import datetime, time as dt_time
import time
import gspread
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook, Workbook
import os

# === Manual start time ===
START_TIME = dt_time(hour=15, minute=17)
print(f"Attendance scheduled for: {START_TIME.strftime('%H:%M')}")

# === Image paths of students ===
image_paths = {
    "Anusha-80": r"C:/Users/anush/Downloads/Minee/anusha80.png",
    "Sahana-144": r"C:/Users/anush/Downloads/Minee/sahana144.png",
    "Anusha-73": r"C:/Users/anush/Downloads/Minee/anusha73.png",
}

# === Google Sheet Setup ===
SERVICE_ACCOUNT_FILE = r"C:/Users/anush/Downloads/Minee/credentials.json"
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
credentials = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
gc = gspread.authorize(credentials)
SPREADSHEET_ID = "1j5cxov8g0jl4Ou6M2ehzcwA-MPBXO8pn85nHTCHFqAg"
spreadsheet = gc.open_by_key(SPREADSHEET_ID)

# === Local Excel file ===
EXCEL_PATH = "C:/Users/anush/Downloads/SmartAttendanceWeb-main/attend.xlsx"

# === Load known faces ===
known_face_encodings = []
known_face_names = []

for name, path in image_paths.items():
    try:
        image = face_recognition.load_image_file(path)
        face_encoding = face_recognition.face_encodings(image)[0]
        known_face_encodings.append(face_encoding)
        known_face_names.append(name)
    except Exception as e:
        print(f"Error loading image for {name}: {e}")

def update_google_sheet(sheet, date_header, name):
    try:
        headers = sheet.row_values(1)
        if not headers or headers[0] != "Name":
            sheet.update_cell(1, 1, "Name")
            headers.insert(0, "Name")

        row_values = sheet.col_values(1)
        if len(row_values) == 1:
            for i, student in enumerate(known_face_names, start=2):
                sheet.update_cell(i, 1, student)

        row_values = sheet.col_values(1)
        if date_header not in headers:
            sheet.update_cell(1, len(headers) + 1, date_header)
            headers.append(date_header)

        col_index = headers.index(date_header) + 1

        for student in known_face_names:
            row_index = row_values.index(student) + 1 if student in row_values else len(row_values) + 1
            current_value = sheet.cell(row_index, col_index).value
            if student == name:
                sheet.update_cell(row_index, col_index, "Present")
            elif not current_value:
                sheet.update_cell(row_index, col_index, "Absent")
    except Exception as e:
        print(f"Error updating Google Sheet: {e}")

def update_excel(date_header, name):
    try:
        if os.path.exists(EXCEL_PATH):
            workbook = load_workbook(EXCEL_PATH)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.cell(row=1, column=1, value="Name")

        sheet.cell(row=1, column=1, value="Name")
        existing_names = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1)]

        for student in known_face_names:
            if student not in existing_names:
                sheet.append([student])

        names_in_sheet = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1)]
        headers = [sheet.cell(row=1, column=j).value for j in range(1, sheet.max_column + 1)]

        if date_header not in headers:
            col_index = len(headers) + 1
            sheet.cell(row=1, column=col_index, value=date_header)
        else:
            col_index = headers.index(date_header) + 1

        for i, student in enumerate(names_in_sheet, start=2):
            current_val = sheet.cell(row=i, column=col_index).value
            if student == name:
                sheet.cell(row=i, column=col_index, value="Present")
            elif current_val is None:
                sheet.cell(row=i, column=col_index, value="Absent")

        workbook.save(EXCEL_PATH)
    except Exception as e:
        print(f"Error updating Excel sheet: {e}")

def take_attendance():
    print("📸 Starting Face Recognition...")
    date_header = datetime.now().strftime("%Y-%m-%d")

    # Ensure worksheet
    worksheets = spreadsheet.worksheets()
    sheet_titles = [ws.title for ws in worksheets]
    if "Attendance" not in sheet_titles:
        spreadsheet.add_worksheet(title="Attendance", rows="100", cols="10")
    sheet = spreadsheet.worksheet("Attendance")

    capture = cv2.VideoCapture(0)
    if not capture.isOpened():
        print("🚫 Could not access the camera.")
        time.sleep(2)
        return

    already_marked = set()
    start_time = time.time()

    while time.time() - start_time < 120:  # 2 minutes
        ret, frame = capture.read()
        if not ret:
            print("⚠️ Frame read error.")
            break

        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb_frame)
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)

        for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
            name = "Unknown"
            distances = face_recognition.face_distance(known_face_encodings, face_encoding)
            best_match = np.argmin(distances)
            if distances[best_match] < 0.6:
                name = known_face_names[best_match]

            if name != "Unknown" and name not in already_marked:
                print(f"✅ Recognized: {name}")
                already_marked.add(name)
                update_google_sheet(sheet, date_header, name)
                update_excel(date_header, name)

            cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
            cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 255, 0), cv2.FILLED)
            font = cv2.FONT_HERSHEY_DUPLEX
            cv2.putText(frame, name, (left + 6, bottom - 6), font, 0.7, (255, 255, 255), 1)

        cv2.imshow("Attendance System", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    capture.release()
    cv2.destroyAllWindows()
    print("🛑 Attendance session ended.")

# === Auto-start logic ===
start_wait = time.time()
while True:
    current_time = datetime.now().time()
    print(f"⏳ Waiting for start time: {START_TIME.strftime('%H:%M')}")
    if current_time.hour == START_TIME.hour and current_time.minute == START_TIME.minute:
        print("🕒 Time matched — taking attendance...")
        take_attendance()
        break

    if time.time() - start_wait > 120:
        print("⏰ Timeout reached. No attendance taken.")
        break

    time.sleep(10)
